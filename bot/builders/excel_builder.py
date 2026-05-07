import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")
CURRENCY_COLS_HINT = {"revenue", "total", "avg_order_value", "price", "amount"}

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def _thin_border():
    side = Side(style="thin", color="D9D9D9")
    return Border(left=side, right=side, top=side, bottom=side)


class ExcelBuilder:
    def __init__(self, config: dict):
        self.config = config

    def build(self, df: pd.DataFrame, output_path: str) -> str:
        wb = Workbook()
        wb.remove(wb.active)

        self._add_data_sheet(wb, df)
        self._add_summary_sheet(wb, df)
        self._add_chart_sheet(wb, df)
        self._add_info_sheet(wb, df)

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        logger.info("Excel report saved to %s", output_path)
        return output_path

    # ── Data sheet ──────────────────────────────────────────────────────────

    def _add_data_sheet(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("Data")

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    if r_idx % 2 == 0:
                        cell.fill = ALT_ROW_FILL
                    cell.alignment = Alignment(vertical="center")

                cell.border = _thin_border()

        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Freeze header
        ws.freeze_panes = "A2"

        # Number formats
        for c_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(c_idx)
            if any(hint in col_name.lower() for hint in CURRENCY_COLS_HINT):
                for row in ws.iter_rows(min_row=2, min_col=c_idx, max_col=c_idx):
                    for cell in row:
                        cell.number_format = '"$"#,##0.00'
            if "date" in col_name.lower():
                for row in ws.iter_rows(min_row=2, min_col=c_idx, max_col=c_idx):
                    for cell in row:
                        cell.number_format = "YYYY-MM-DD"
            if "pct" in col_name.lower() or "rate" in col_name.lower():
                for row in ws.iter_rows(min_row=2, min_col=c_idx, max_col=c_idx):
                    for cell in row:
                        cell.number_format = "0.00%"

        # Conditional formatting: negatives red, positives green
        num_cols = df.select_dtypes(include="number").columns.tolist()
        for col_name in num_cols:
            c_idx = df.columns.tolist().index(col_name) + 1
            col_letter = get_column_letter(c_idx)
            rng = f"{col_letter}2:{col_letter}{len(df) + 1}"
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
            )
            ws.conditional_formatting.add(
                rng,
                CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL),
            )

    # ── Summary sheet ────────────────────────────────────────────────────────

    def _add_summary_sheet(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("Summary")
        num_df = df.select_dtypes(include="number")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        title = ws.cell(row=1, column=1, value=self.config.get("name", "Report Summary"))
        title.font = Font(bold=True, size=14, color="1F3864")
        title.alignment = Alignment(horizontal="left")

        ws.cell(row=2, column=1, value="Generated").font = Font(italic=True, color="888888")
        ws.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

        row = 4
        ws.cell(row=row, column=1, value="Metric").fill = HEADER_FILL
        ws.cell(row=row, column=1).font = HEADER_FONT
        ws.cell(row=row, column=2, value="Value").fill = HEADER_FILL
        ws.cell(row=row, column=2).font = HEADER_FONT

        row += 1
        ws.cell(row=row, column=1, value="Total rows")
        ws.cell(row=row, column=2, value=len(df))
        row += 1

        for col in num_df.columns:
            ws.cell(row=row, column=1, value=f"Total {col}")
            ws.cell(row=row, column=2, value=round(num_df[col].sum(), 2))
            row += 1
            ws.cell(row=row, column=1, value=f"Average {col}")
            ws.cell(row=row, column=2, value=round(num_df[col].mean(), 2))
            row += 1
            ws.cell(row=row, column=1, value=f"Max {col}")
            ws.cell(row=row, column=2, value=round(num_df[col].max(), 2))
            row += 1

    # ── Chart sheet ──────────────────────────────────────────────────────────

    def _add_chart_sheet(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("Chart")
        num_cols = df.select_dtypes(include="number").columns.tolist()

        if not num_cols:
            ws.cell(row=1, column=1, value="No numeric data for chart")
            return

        # Write a minimal table for the chart
        col_name = num_cols[0]
        ws.cell(row=1, column=1, value="Row")
        ws.cell(row=1, column=2, value=col_name)
        for i, val in enumerate(df[col_name].head(50), start=2):
            ws.cell(row=i, column=1, value=i - 1)
            ws.cell(row=i, column=2, value=val)

        n = min(len(df), 50) + 1
        chart_type = self.config.get("chart_type", "bar")
        chart = BarChart() if chart_type == "bar" else LineChart()
        chart.title = self.config.get("name", "Report Chart")
        chart.style = 10
        chart.y_axis.title = col_name
        chart.x_axis.title = "Row"

        data = Reference(ws, min_col=2, min_row=1, max_row=n)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, "D2")

    # ── Info sheet ───────────────────────────────────────────────────────────

    def _add_info_sheet(self, wb: Workbook, df: pd.DataFrame):
        ws = wb.create_sheet("Info")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 40

        info = [
            ("Report Name", self.config.get("name", "—")),
            ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Source Type", self.config.get("source", {}).get("type", "—")),
            ("Row Count", len(df)),
            ("Column Count", len(df.columns)),
            ("Columns", ", ".join(df.columns.tolist())),
            ("Schedule", self.config.get("schedule", "—")),
        ]

        ws.cell(row=1, column=1, value="Report Metadata").font = Font(bold=True, size=13, color="1F3864")

        for r, (key, val) in enumerate(info, start=3):
            ws.cell(row=r, column=1, value=key).font = Font(bold=True)
            ws.cell(row=r, column=2, value=str(val))
