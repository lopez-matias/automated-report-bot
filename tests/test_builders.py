import os
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from bot.builders.excel_builder import ExcelBuilder
from bot.builders.pdf_builder import PdfBuilder


@pytest.fixture
def sample_df():
    return pd.DataFrame(
        {
            "date": ["2025-01-01", "2025-01-02", "2025-01-03", "2025-01-04"],
            "region": ["East", "West", "North", "South"],
            "revenue": [1000.0, 1500.0, 800.0, 2200.0],
            "orders": [10, 15, 8, 22],
            "avg_order_value": [100.0, 100.0, 100.0, 100.0],
        }
    )


@pytest.fixture
def report_config():
    return {
        "name": "Test Report",
        "schedule": "every monday at 08:00",
        "source": {"type": "csv"},
        "chart_type": "bar",
    }


# ── Excel Builder ────────────────────────────────────────────────────────────

class TestExcelBuilder:
    def test_creates_file(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.xlsx")
        builder = ExcelBuilder(report_config)
        result = builder.build(sample_df, out)
        assert Path(result).exists()

    def test_has_four_sheets(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.xlsx")
        ExcelBuilder(report_config).build(sample_df, out)
        wb = load_workbook(out)
        assert set(wb.sheetnames) == {"Data", "Summary", "Chart", "Info"}

    def test_data_sheet_row_count(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.xlsx")
        ExcelBuilder(report_config).build(sample_df, out)
        wb = load_workbook(out)
        ws = wb["Data"]
        # 1 header + 4 data rows
        assert ws.max_row == 5

    def test_header_fill_is_dark(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.xlsx")
        ExcelBuilder(report_config).build(sample_df, out)
        wb = load_workbook(out)
        ws = wb["Data"]
        # Header cell should have fill (not None / empty)
        fill_type = ws["A1"].fill.fill_type
        assert fill_type == "solid"

    def test_info_sheet_has_report_name(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.xlsx")
        ExcelBuilder(report_config).build(sample_df, out)
        wb = load_workbook(out)
        ws = wb["Info"]
        values = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
        assert "Test Report" in values

    def test_empty_dataframe(self, tmp_path, report_config):
        df = pd.DataFrame({"a": [], "b": []})
        out = str(tmp_path / "empty.xlsx")
        ExcelBuilder(report_config).build(df, out)
        assert Path(out).exists()


# ── PDF Builder ──────────────────────────────────────────────────────────────

class TestPdfBuilder:
    def test_creates_file(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.pdf")
        builder = PdfBuilder(report_config)
        result = builder.build(sample_df, out)
        assert Path(result).exists()

    def test_file_is_pdf(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.pdf")
        PdfBuilder(report_config).build(sample_df, out)
        with open(out, "rb") as f:
            header = f.read(4)
        assert header == b"%PDF"

    def test_file_nonzero_size(self, tmp_path, sample_df, report_config):
        out = str(tmp_path / "report.pdf")
        PdfBuilder(report_config).build(sample_df, out)
        assert Path(out).stat().st_size > 1024

    def test_large_dataframe(self, tmp_path, report_config):
        df = pd.DataFrame({"col1": range(300), "col2": range(300, 600)})
        out = str(tmp_path / "large.pdf")
        PdfBuilder(report_config).build(df, out)
        assert Path(out).exists()
